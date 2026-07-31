"""Parses a user-uploaded CSV/Excel file of test cases into the same raw
row shape `app.indexer.graph.testrail_builder.build_graph` already
consumes for a real TestRail sync — see that module's `TestRailProjectData`
docstring for why the two sources share one builder.

Deliberately tolerant of header naming: a spreadsheet someone already has
lying around was very likely not written with this tool's exact column
names in mind, so this matches on a few common synonyms per field rather
than requiring one exact schema.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import AppError

# Bounds mirror TestRailTool's own _MAX_ITEMS - a spreadsheet with more
# rows than this is almost certainly the wrong file, not a legitimately
# huge test suite, and parsing it fully would block the request for too
# long (this endpoint is synchronous, not a background job - see
# app.services.test_case_upload_service's own docstring for why that's
# fine at this scale).
_MAX_ROWS = 5_000
_MAX_FILE_BYTES = 5 * 1024 * 1024

_TITLE_HEADERS = {"title", "test case", "test case title", "name", "summary"}
_SECTION_HEADERS = {"section", "suite", "module", "category", "group"}
_PRIORITY_HEADERS = {"priority"}
_REFS_HEADERS = {"refs", "reference", "references", "jira", "ticket", "requirement"}


class TestCaseUploadParseError(AppError):
    status_code = 422
    error_code = "test_case_upload_parse_error"


def _match_header(headers: list[str], candidates: set[str]) -> str | None:
    for header in headers:
        if header.strip().lower() in candidates:
            return header
    return None


def _rows_to_cases(headers: list[str], rows: list[list[Any]]) -> list[dict[str, str]]:
    if not headers:
        raise TestCaseUploadParseError("The file has no header row.")

    title_col = _match_header(headers, _TITLE_HEADERS)
    section_col = _match_header(headers, _SECTION_HEADERS)
    priority_col = _match_header(headers, _PRIORITY_HEADERS)
    refs_col = _match_header(headers, _REFS_HEADERS)
    # No recognized title column: assume the simplest common case, a
    # single column of test case titles with an arbitrary header (or
    # none) - the first column, whatever it's called.
    title_index = headers.index(title_col) if title_col else 0
    section_index = headers.index(section_col) if section_col else None
    priority_index = headers.index(priority_col) if priority_col else None
    refs_index = headers.index(refs_col) if refs_col else None

    def cell(row: list[Any], index: int | None) -> str:
        if index is None or index >= len(row) or row[index] is None:
            return ""
        return str(row[index]).strip()

    cases: list[dict[str, str]] = []
    for row in rows:
        title = cell(row, title_index)
        if not title:
            continue  # a blank row, or a row with no title - not a case
        cases.append(
            {
                "title": title,
                "section": cell(row, section_index) or "Uploaded Test Cases",
                "priority": cell(row, priority_index),
                "refs": cell(row, refs_index),
            }
        )
    return cases


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    return _rows_to_cases(rows[0], rows[1 : _MAX_ROWS + 1])


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise TestCaseUploadParseError(f"Couldn't read this Excel file: {exc}") from exc
    sheet = workbook.worksheets[0]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return _rows_to_cases(headers, rows[1 : _MAX_ROWS + 1])


def parse_test_case_file(filename: str, content: bytes) -> list[dict[str, str]]:
    """Returns one dict per test case: `{"title", "section", "priority",
    "refs"}` — the same shape `testrail_builder.build_graph` expects for
    its `cases` list (via `TestCaseUploadService`, which assigns each a
    sequential integer id and groups them into sections before handing
    off to that builder)."""
    if len(content) > _MAX_FILE_BYTES:
        raise TestCaseUploadParseError(
            f"File is too large ({len(content) // 1024} KB) — the limit is "
            f"{_MAX_FILE_BYTES // 1024} KB."
        )

    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if suffix == "csv":
        cases = _parse_csv(content)
    elif suffix in ("xlsx", "xlsm"):
        cases = _parse_xlsx(content)
    elif suffix == "xls":
        raise TestCaseUploadParseError(
            "The legacy .xls format isn't supported — save this file as .xlsx or .csv."
        )
    else:
        raise TestCaseUploadParseError(
            f"Unsupported file type '.{suffix}' — upload a .csv or .xlsx file."
        )

    if not cases:
        raise TestCaseUploadParseError("No test cases found in this file.")
    return cases
